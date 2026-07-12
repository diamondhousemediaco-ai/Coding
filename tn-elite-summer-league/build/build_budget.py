#!/usr/bin/env python3
"""Build the tiered league budget workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WB = openpyxl.Workbook()

ARIAL = "Arial"
BLUE = Font(name=ARIAL, size=10, color="0000FF")           # hardcoded inputs
BLACK = Font(name=ARIAL, size=10)
GREEN = Font(name=ARIAL, size=10, color="008000")          # cross-sheet links
BOLD = Font(name=ARIAL, size=10, bold=True)
BOLD12 = Font(name=ARIAL, size=12, bold=True)
TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="1F3864")
SECFILL = PatternFill("solid", fgColor="D9E2F2")
YELLOW = PatternFill("solid", fgColor="FFFF00")
TOTFILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))
CUR = '$#,##0;($#,##0);-'
NUM = '#,##0;(#,##0);-'

def sheet_title(ws, text, span=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE; c.fill = HDRFILL
    for col in range(1, span+1):
        ws.cell(row=1, column=col).fill = HDRFILL
    ws.row_dimensions[1].height = 24

def section(ws, row, text, span=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD; c.fill = SECFILL
    for col in range(1, span+1):
        ws.cell(row=row, column=col).fill = SECFILL
    return row + 1

# ---------------------------------------------------------------- READ ME
rm = WB.active
rm.title = "READ ME"
sheet_title(rm, "Tennessee Elite Summer Basketball League — Budget Model", 4)
rows = [
    ("", ""),
    ("What this file is", "A working budget model for a 16-team (8 boys / 8 girls) elite HS 5v5 summer league at Belmont University, first season Summer 2027. Two tiers: a Year 1 launch budget and a Full Vision (Year 3) budget."),
    ("", ""),
    ("How to use it", "Every number in BLUE on the Assumptions and Staffing Plan tabs is an editable input. Everything in black is a formula that recalculates. Yellow cells are the assumptions that most need real quotes (Belmont rental, sponsor commitments)."),
    ("", ""),
    ("Color legend", "BLUE = editable input · BLACK = formula · GREEN = pulls from another tab · YELLOW fill = key assumption, verify with a real quote"),
    ("", ""),
    ("Tabs", "Assumptions → all levers. Year 1 Budget → launch season. Full Vision Budget → the ESPN-scale version. Revenue Scenarios → conservative / base / stretch. Sponsor Menu → sellable assets with price points. Staffing Plan → the team that runs it, with comp."),
    ("", ""),
    ("Key sources", "Cost benchmarks: ZebraWeb officials pay survey; ArenaMediaCloud production calculator; Prep Athletic Director sponsorship pricing method; NCAA ECAG fee schedule; TSSAA 2024-25 Handbook & basketball calendar; Unrivaled/OTE reporting (ESPN, SportsPro, Front Office Sports). Venue costs are ESTIMATES — Belmont does not publish rates (contact osves@belmont.edu)."),
    ("", ""),
    ("Important caveat", "This model is planning-grade, not accounting-grade. Every venue, production, and sponsorship figure should be replaced with signed quotes as they arrive."),
]
r = 3
for label, text in rows:
    rm.cell(row=r, column=1, value=label).font = BOLD
    c = rm.cell(row=r, column=2, value=text)
    c.font = BLACK; c.alignment = Alignment(wrap_text=True, vertical="top")
    if text: rm.row_dimensions[r].height = max(15, 13 * (len(text)//95 + 1))
    r += 1
rm.column_dimensions['A'].width = 20
rm.column_dimensions['B'].width = 110

# ---------------------------------------------------------------- ASSUMPTIONS
A = WB.create_sheet("Assumptions")
sheet_title(A, "Assumptions — every blue cell is editable", 4)
A.column_dimensions['A'].width = 46
A.column_dimensions['B'].width = 14
A.column_dimensions['C'].width = 12
A.column_dimensions['D'].width = 95

ROW = {}
def put(ws, r, label, value, unit="", note="", fmt=NUM, formula=False, key=False):
    ws.cell(row=r, column=1, value=label).font = BLACK
    c = ws.cell(row=r, column=2, value=value)
    c.font = BLACK if formula else BLUE
    c.number_format = fmt
    if key: c.fill = YELLOW
    ws.cell(row=r, column=3, value=unit).font = Font(name=ARIAL, size=9, italic=True)
    n = ws.cell(row=r, column=4, value=note)
    n.font = Font(name=ARIAL, size=9, color="595959")
    for col in range(1,5): ws.cell(row=r, column=col).border = THIN
    return r

r = 3
r = section(A, r, "LEAGUE STRUCTURE", 4)
ROW['teams_gender']   = put(A, r, "Teams per gender", 8); r+=1
ROW['genders']        = put(A, r, "Genders (boys + girls)", 2); r+=1
ROW['teams_total']    = put(A, r, "Total teams", f"=B{ROW['teams_gender']}*B{ROW['genders']}", formula=True); r+=1
ROW['roster']         = put(A, r, "Roster size per team", 12); r+=1
ROW['players']        = put(A, r, "Total players", f"=B{ROW['teams_total']}*B{ROW['roster']}", formula=True); r+=1
ROW['sessions']       = put(A, r, "Sessions per summer", 2, note="Two 5-week sessions scheduled around the TSSAA dead period (~Jun 21 - Jul 4, 2027)"); r+=1
ROW['gpt']            = put(A, r, "Regular-season games per team per session", 8, note="2 games/week x 4 weeks; week 5 = playoffs"); r+=1
ROW['playoff_g']      = put(A, r, "Playoff games per gender per session", 7, note="8-team single-elim + 3rd place + final"); r+=1
ROW['games_session']  = put(A, r, "Games per session (both genders)", f"=(B{ROW['teams_gender']}*B{ROW['gpt']}/2+B{ROW['playoff_g']})*B{ROW['genders']}", formula=True); r+=1
ROW['games_reg']      = put(A, r, "Total league games per summer", f"=B{ROW['games_session']}*B{ROW['sessions']}", formula=True); r+=1
ROW['games_natl_y1']  = put(A, r, "Nationals invitational games — Year 1", 0, note="Recommend launching Nationals in Year 2-3, not Year 1"); r+=1
ROW['games_natl_fv']  = put(A, r, "Nationals invitational games — Full Vision", 24, note="16 invited teams (8B/8G), ~10-day event aligned to NCAA July live periods"); r+=1
ROW['event_days_y1']  = put(A, r, "Game/event days — Year 1", 44, note="~20 game days/session + combine, draft, championship days"); r+=1
ROW['event_days_fv']  = put(A, r, "Game/event days — Full Vision", 64, note="Adds Nationals (~10 days) and more arena dates"); r+=1
r += 1

r = section(A, r, "PLAYER FEES (your hybrid model)", 4)
ROW['tryouts']        = put(A, r, "Tryout participants", 600, note="Open tryouts across TN; fee credited toward season fee if drafted"); r+=1
ROW['tryout_fee']     = put(A, r, "Tryout fee", 50, fmt=CUR, note="Rolled into season fee for players who make it; combine fee waived for invitees"); r+=1
ROW['player_fee_y1']  = put(A, r, "Player season fee — Year 1 (net of tryout credit)", 400, fmt=CUR, key=True, note="All-in for both sessions. vs $3,000-5,000+ typical AAU summer. Full Vision goal = $0 (fully sponsored)"); r+=1
ROW['player_fee_fv']  = put(A, r, "Player season fee — Full Vision", 0, fmt=CUR, note="The anti-AAU endgame: sponsors fund everything"); r+=1
r += 1

r = section(A, r, "COST RATES — shared drivers", 4)
ROW['ref_rate']       = put(A, r, "Officials — per official per game", 95, fmt=CUR, note="Elite HS varsity level runs $70-120/game (ZebraWeb 2025); 3-person certified crews"); r+=1
ROW['refs_per_game']  = put(A, r, "Officials per game", 3); r+=1
ROW['trainer_day']    = put(A, r, "Certified athletic trainer — per event day", 450, fmt=CUR, note="Per-diem ATs ~$40-75/hr; required by NCAA ECAG at every facility"); r+=1
ROW['opsday_y1']      = put(A, r, "Game-day ops crew per day — Year 1", 800, fmt=CUR, note="Scorekeepers, clock, PA/DJ, gate, security"); r+=1
ROW['opsday_fv']      = put(A, r, "Game-day ops crew per day — Full Vision", 1400, fmt=CUR); r+=1
r += 1

r = section(A, r, "VENUE (estimates — get Belmont quote: osves@belmont.edu)", 4)
ROW['pf_hours_y1']    = put(A, r, "Belmont practice facility court-hours — Year 1", 320, note="2 courts, ~16 hrs/wk across 10 game weeks; 45,000 sq ft facility opened 2021, two full courts"); r+=1
ROW['pf_rate']        = put(A, r, "Practice facility rate per court-hour", 90, fmt=CUR, key=True, note="Market: $50-150/hr for quality indoor courts; university premium assumed"); r+=1
ROW['arena_days_y1']  = put(A, r, "Curb Event Center days — Year 1", 6, note="Draft night + marquee weekends + championships. 5,085-seat arena"); r+=1
ROW['arena_days_fv']  = put(A, r, "Curb Event Center days — Full Vision", 20, note="All game days in the arena, Unrivaled-style single-site look"); r+=1
ROW['arena_rate']     = put(A, r, "Curb Event Center per event day", 7500, fmt=CUR, key=True, note="ESTIMATE. Comparable college arenas ~$2,500-10,000+/day plus staffing. Belmont does not publish rates"); r+=1
ROW['combine_venue']  = put(A, r, "Combine + draft venue & ops", 30000, fmt=CUR, note="2-day combine (testing, measurements, scrimmages) + televised-style draft night"); r+=1
r += 1

r = section(A, r, "PRODUCTION & STREAMING", 4)
ROW['marquee_y1']     = put(A, r, "Marquee broadcast games — Year 1", 16, note="3-4 cameras, commentary, replay, graphics"); r+=1
ROW['marquee_fv']     = put(A, r, "Marquee broadcast games — Full Vision", 60); r+=1
ROW['marquee_rate']   = put(A, r, "Cost per marquee game — Year 1", 2800, fmt=CUR, key=True, note="Clean 3-cam streamed game w/ commentary & graphics = $2,500-6,000 (ArenaMediaCloud/DJC)"); r+=1
ROW['marquee_rate_fv']= put(A, r, "Cost per marquee game — Full Vision", 6000, fmt=CUR, note="8+ cam broadcast tier"); r+=1
ROW['std_rate']       = put(A, r, "Cost per standard-stream game", 400, fmt=CUR, note="Auto-cam (Pixellot/BallerTV-style) + student commentary; near-zero marginal cost after install"); r+=1
ROW['std_rate_fv']    = put(A, r, "Cost per standard-stream game — Full Vision", 1500, fmt=CUR); r+=1
r += 1

r = section(A, r, "PLAYER EXPERIENCE & GEAR", 4)
ROW['kit_y1']         = put(A, r, "Uniform/gear package per player — Year 1", 220, fmt=CUR, note="Sublimated home/away + shooting shirt + bag; premium packages run $120-250"); r+=1
ROW['kit_fv']         = put(A, r, "Uniform/gear package per player — Full Vision", 350, fmt=CUR, note="True Nike/adidas team-catalog tier"); r+=1
ROW['equip']          = put(A, r, "Balls & court equipment", 6000, fmt=CUR, note="~24 Wilson Evo NXT game balls (~$110 ea) + practice balls, carts, boards"); r+=1
ROW['px_fv']          = put(A, r, "Player meals/transport fund — Full Vision", 80000, fmt=CUR); r+=1
ROW['nil_fv']         = put(A, r, "Sponsor-funded player NIL/appearance pool — Full Vision", 100000, fmt=CUR, note="TSSAA allows HS NIL (Dec 2022, no cap) — no school marks, no coach facilitation"); r+=1
r += 1

r = section(A, r, "OTHER FIXED COSTS", 4)
ROW['mktg_y1']        = put(A, r, "Marketing & brand — Year 1", 45000, fmt=CUR, note="Brand system, content creators, launch events, paid social, PR — the Unrivaled growth engine"); r+=1
ROW['mktg_fv']        = put(A, r, "Marketing & brand — Full Vision", 150000, fmt=CUR); r+=1
ROW['ins_y1']         = put(A, r, "Insurance — Year 1", 18000, fmt=CUR, note="CGL $1M/$2M + accident medical, NCAA named additional insured (ECAG requirement)"); r+=1
ROW['ins_fv']         = put(A, r, "Insurance — Full Vision", 35000, fmt=CUR); r+=1
ROW['compliance']     = put(A, r, "NCAA cert, USAB Gold licenses, background checks", 6000, fmt=CUR, note="ECAG: $30 operator + $50 app + $350 event; USAB Gold ~$68/coach x ~40; checks ~$15/staff"); r+=1
ROW['tech_y1']        = put(A, r, "Tech & stats — Year 1", 15000, fmt=CUR, note="Stats platform, website, registration, highlight tooling — stats packages are a recruiting differentiator"); r+=1
ROW['tech_fv']        = put(A, r, "Tech & stats — Full Vision", 40000, fmt=CUR); r+=1
ROW['legal_y1']       = put(A, r, "Legal, entity & accounting — Year 1", 12000, fmt=CUR, note="TN nonprofit charter $100 + IRS 1023 $600 + charitable reg $50 + counsel/CPA"); r+=1
ROW['legal_fv']       = put(A, r, "Legal & accounting — Full Vision", 30000, fmt=CUR); r+=1
ROW['awards_y1']      = put(A, r, "Awards & trophies — Year 1", 6000, fmt=CUR); r+=1
ROW['awards_fv']      = put(A, r, "Awards & trophies — Full Vision", 15000, fmt=CUR); r+=1
ROW['misc_y1']        = put(A, r, "Travel & hospitality — Year 1", 8000, fmt=CUR); r+=1
ROW['natl_travel_fv'] = put(A, r, "Nationals travel subsidies — Full Vision", 120000, fmt=CUR, note="16 visiting teams x ~$7,500 toward travel/housing"); r+=1
ROW['natl_ops_fv']    = put(A, r, "Nationals event ops — Full Vision", 60000, fmt=CUR); r+=1
ROW['conting']        = put(A, r, "Contingency %", 0.10, fmt='0.0%', key=True, note="Stored as a fraction"); r+=1
r += 1

r = section(A, r, "REVENUE DRIVERS — Year 1 base case", 4)
ROW['title_y1']       = put(A, r, "Title sponsor", 50000, fmt=CUR, key=True, note="Regional property w/ streamed games: title runs $10K-50K (Prep AD method + ZipSprout data)"); r+=1
ROW['pres_n_y1']      = put(A, r, "Presenting sponsors — count", 2); r+=1
ROW['pres_p_y1']      = put(A, r, "Presenting sponsors — price", 25000, fmt=CUR); r+=1
ROW['cat_n_y1']       = put(A, r, "Official category partners — count", 5, note="Ball, apparel, drink, recovery, bank etc. Category exclusivity commands 2-5x premium"); r+=1
ROW['cat_p_y1']       = put(A, r, "Official category partners — price", 10000, fmt=CUR); r+=1
ROW['jers_n_y1']      = put(A, r, "Team jersey partners — count", 12, note="One local business per team (16 available)"); r+=1
ROW['jers_p_y1']      = put(A, r, "Team jersey partners — price", 5000, fmt=CUR); r+=1
ROW['gate_y1']        = put(A, r, "Gate revenue — Year 1", 120000, fmt=CUR, key=True, note="6 arena days x ~1,500 x $12 + regular nights x ~250 x $8. Replace with ticket model as dates firm up"); r+=1
ROW['stream_y1']      = put(A, r, "Streaming rev share / league pass — Year 1", 15000, fmt=CUR, note="BallerTV-style parent subscriptions rev share"); r+=1
ROW['merch_y1']       = put(A, r, "Merchandise net — Year 1", 20000, fmt=CUR); r+=1
ROW['grants_y1']      = put(A, r, "Donations & grants (501c3) — Year 1", 60000, fmt=CUR, note="Community foundations, youth-sports grants, individual donors"); r+=1
r += 1

r = section(A, r, "REVENUE DRIVERS — Full Vision", 4)
ROW['title_fv']       = put(A, r, "Title sponsor — Full Vision", 250000, fmt=CUR, note="Between regional ($50K) and minor-league pro ($500K) tiers once streamed + televised"); r+=1
ROW['spons_other_fv'] = put(A, r, "All other sponsorship — Full Vision", 750000, fmt=CUR, note="Presenting, category, jersey, court, content series — see Sponsor Menu"); r+=1
ROW['gate_fv']        = put(A, r, "Gate revenue — Full Vision", 500000, fmt=CUR, note="20 arena days x ~1,800 avg x $14"); r+=1
ROW['media_fv']       = put(A, r, "Media/streaming revenue — Full Vision", 150000, fmt=CUR, note="Time-buy/barter path to ESPN+ via packagers like Paragon; league pass"); r+=1
ROW['merch_fv']       = put(A, r, "Merchandise net — Full Vision", 60000, fmt=CUR); r+=1
ROW['natl_entry_fv']  = put(A, r, "Nationals team entry fees", 80000, fmt=CUR, note="16 invited teams x $5K (TBT charges $125K-250K at pro level)"); r+=1
ROW['grants_fv']      = put(A, r, "Donations & grants — Full Vision", 100000, fmt=CUR); r+=1

AS = "Assumptions"
def aref(key): return f"={AS}!$B${ROW[key]}"
def a(key): return f"{AS}!$B${ROW[key]}"

# ---------------------------------------------------------------- STAFFING PLAN
S = WB.create_sheet("Staffing Plan")
sheet_title(S, "Staffing Plan — the team that runs everything", 6)
for col, w in zip("ABCDEF", [40, 16, 16, 16, 16, 70]):
    S.column_dimensions[col].width = w
hdr = ["Role", "Type — Year 1", "Year 1 Comp", "Type — Full Vision", "Full Vision Comp", "What they own"]
r = 3
for i, h in enumerate(hdr, 1):
    c = S.cell(row=r, column=i, value=h); c.font = BOLD; c.fill = SECFILL
r += 1
staff_rows = []
def staff(r, role, t1, c1, t2, c2, owns):
    S.cell(row=r, column=1, value=role).font = BLACK
    S.cell(row=r, column=2, value=t1).font = BLACK
    x = S.cell(row=r, column=3, value=c1); x.font = BLUE; x.number_format = CUR
    S.cell(row=r, column=4, value=t2).font = BLACK
    y = S.cell(row=r, column=5, value=c2); y.font = BLUE; y.number_format = CUR
    d = S.cell(row=r, column=6, value=owns); d.font = Font(name=ARIAL, size=9, color="595959")
    for col in range(1,7): S.cell(row=r, column=col).border = THIN
    staff_rows.append(r)
    return r+1

r = section(S, r, "FRONT OFFICE (league staff)", 6)
r = staff(r, "Commissioner / League President (founder)", "Founder stipend", 25000, "Full-time", 120000, "Vision, sponsor relationships, media deals, public face of the league")
r = staff(r, "Director of Basketball Operations", "Seasonal contract", 30000, "Full-time", 90000, "Combine, draft, schedules, officials, rosters, competition rules")
r = staff(r, "Director of Partnerships & Revenue", "Seasonal + commission", 25000, "Full-time + commission", 90000, "Sponsor sales & activation, ticketing, renewals. Add 10% commission on new sponsor revenue")
r = staff(r, "Director of Media & Content", "Seasonal contract", 30000, "Full-time", 85000, "Streaming, social, brand, highlight engine — the Unrivaled-style growth lever")
r = staff(r, "Director of Events & Logistics", "Seasonal contract", 25000, "Full-time", 75000, "Venue, game-day ops, vendors, security, medical coverage")
r = staff(r, "Compliance & Player Welfare Coordinator", "Part-time", 12000, "Full-time", 65000, "TSSAA/NCAA ECAG compliance, eligibility, SafeSport, background checks, parent comms")
r = staff(r, "Finance & Admin", "Fractional bookkeeper", 8000, "Full-time ops manager", 65000, "Books, payroll, registrations, insurance")
r = staff(r, "Coordinators & interns (Belmont/Lipscomb/TSU sport mgmt)", "Stipends", 6000, "4-6 paid coordinators", 120000, "Social clips, stats, team services, gameday support — recruit from local sport management programs")
fo_first, fo_last = staff_rows[0], staff_rows[-1]
r += 1
r = section(S, r, "TEAM STAFF (16 teams: 8 boys, 8 girls)", 6)
r = staff(r, "Head Coaches x16", "Season stipend $3,500 ea", 56000, "$8,000 ea", 128000, "Recruit college assistants, former pros, top trainers — NOT the players' own HS coaches (avoids TSSAA dead-period conflicts)")
r = staff(r, "Assistant Coaches x16", "Season stipend $1,750 ea", 28000, "$4,000 ea", 64000, "One per team; development focus")
r = staff(r, "Interim/floater coach pool (4-6)", "Per-game stipends", 6000, "Per-game stipends", 12000, "League-wide substitutes covering absences — your 'interim coach' insurance policy")
tm_first, tm_last = staff_rows[-3], staff_rows[-1]
r += 1
r = section(S, r, "GAME-DAY & PRODUCTION (variable — costed in the budget tabs, listed here for completeness)", 6)
for role, owns in [
    ("Officials (3-person certified crews)", "Costed per game in budget tabs"),
    ("Certified athletic trainers", "Required at every facility by NCAA ECAG; costed per event day"),
    ("Scorekeepers, clock, PA announcer, DJ, gate, security", "Costed per event day"),
    ("Producer, camera ops, commentators, highlight cutters, photographer", "Costed per game in production lines"),
]:
    S.cell(row=r, column=1, value=role).font = BLACK
    S.cell(row=r, column=6, value=owns).font = Font(name=ARIAL, size=9, color="595959")
    for col in range(1,7): S.cell(row=r, column=col).border = THIN
    r += 1
r += 1
S.cell(row=r, column=1, value="TOTAL — Front office").font = BOLD
c = S.cell(row=r, column=3, value=f"=SUM(C{fo_first}:C{fo_last})"); c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
c = S.cell(row=r, column=5, value=f"=SUM(E{fo_first}:E{fo_last})"); c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
FO_ROW = r
r += 1
S.cell(row=r, column=1, value="TOTAL — Team staff (coaches)").font = BOLD
c = S.cell(row=r, column=3, value=f"=SUM(C{tm_first}:C{tm_last})"); c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
c = S.cell(row=r, column=5, value=f"=SUM(E{tm_first}:E{tm_last})"); c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
TM_ROW = r

SP = "'Staffing Plan'"

# ---------------------------------------------------------------- BUDGET SHEETS
def build_budget(ws_name, year1):
    B = WB.create_sheet(ws_name)
    label = "Year 1 (Summer 2027) — lean but legit" if year1 else "Full Vision (Year 3+) — the Unrivaled-scale version"
    sheet_title(B, f"{ws_name}: {label}", 4)
    B.column_dimensions['A'].width = 48
    B.column_dimensions['B'].width = 16
    B.column_dimensions['C'].width = 4
    B.column_dimensions['D'].width = 85
    r = 3
    r = section(B, r, "EXPENSES", 4)
    exp_first = r
    def line(rr, lbl, formula, note="", link=False):
        B.cell(row=rr, column=1, value=lbl).font = BLACK
        c = B.cell(row=rr, column=2, value=formula)
        c.font = GREEN if link else BLACK
        c.number_format = CUR
        B.cell(row=rr, column=4, value=note).font = Font(name=ARIAL, size=9, color="595959")
        for col in range(1,5): B.cell(row=rr, column=col).border = THIN
        return rr+1

    if year1:
        games = f"{a('games_reg')}+{a('games_natl_y1')}"
        r = line(r, "Venue — practice facility court time", f"={a('pf_hours_y1')}*{a('pf_rate')}", "2 courts at Belmont's 2021 practice facility")
        r = line(r, "Venue — Curb Event Center days", f"={a('arena_days_y1')}*{a('arena_rate')}", "Draft night, marquee weekends, championships (5,085 seats)")
        r = line(r, "Combine + draft events", f"={a('combine_venue')}", "2-day combine with testing/measurements + a produced draft night")
        r = line(r, "Officials", f"=({games})*{a('refs_per_game')}*{a('ref_rate')}", "3-person certified crews, every game")
        r = line(r, "Athletic trainers / medical", f"={a('event_days_y1')}*{a('trainer_day')}", "Certified AT at every event day (NCAA ECAG requirement)")
        r = line(r, "Game-day operations", f"={a('event_days_y1')}*{a('opsday_y1')}", "Scorekeepers, clock, PA/DJ, gate, security")
        r = line(r, "Front office", f"={SP}!$C${FO_ROW}", "See Staffing Plan tab", link=True)
        r = line(r, "Coaches (16 HC + 16 AC + interim pool)", f"={SP}!$C${TM_ROW}", "See Staffing Plan tab", link=True)
        r = line(r, "Uniforms & player gear", f"={a('players')}*{a('kit_y1')}", "Premium sublimated home/away kits + bag")
        r = line(r, "Balls & equipment", f"={a('equip')}", "Wilson Evo NXT game balls + court equipment")
        r = line(r, "Production — marquee broadcast games", f"={a('marquee_y1')}*{a('marquee_rate')}", "3-4 cams, commentary, graphics, replay")
        r = line(r, "Production — standard streams (all other games)", f"=(({games})-{a('marquee_y1')})*{a('std_rate')}", "Auto-cam + student commentary; every game streamed")
        r = line(r, "Marketing & brand", f"={a('mktg_y1')}", "Brand system, content team, launch, paid social, PR")
        r = line(r, "Insurance", f"={a('ins_y1')}", "CGL $1M/$2M + accident medical, NCAA additional insured")
        r = line(r, "Compliance (NCAA cert, USAB licenses, checks)", f"={a('compliance')}")
        r = line(r, "Tech & stats", f"={a('tech_y1')}")
        r = line(r, "Legal, entity & accounting", f"={a('legal_y1')}", "501(c)(3) formation + counsel")
        r = line(r, "Awards & trophies", f"={a('awards_y1')}")
        r = line(r, "Travel & hospitality", f"={a('misc_y1')}")
    else:
        games = f"{a('games_reg')}+{a('games_natl_fv')}"
        r = line(r, "Venue — practice facility court time", f"={a('pf_hours_y1')}*{a('pf_rate')}")
        r = line(r, "Venue — Curb Event Center days", f"={a('arena_days_fv')}*{a('arena_rate')}", "Every game day in the arena — single-site, Unrivaled-style")
        r = line(r, "Combine + draft events", f"={a('combine_venue')}*2.5", "Bigger produced combine + televised-style draft")
        r = line(r, "Officials", f"=({games})*{a('refs_per_game')}*({a('ref_rate')}+15)", "Rate premium for top crews")
        r = line(r, "Athletic trainers / medical", f"={a('event_days_fv')}*({a('trainer_day')}+50)")
        r = line(r, "Game-day operations", f"={a('event_days_fv')}*{a('opsday_fv')}")
        r = line(r, "Front office (full-time)", f"={SP}!$E${FO_ROW}", "See Staffing Plan tab", link=True)
        r = line(r, "Coaches", f"={SP}!$E${TM_ROW}", "See Staffing Plan tab", link=True)
        r = line(r, "Uniforms & player gear", f"={a('players')}*{a('kit_fv')}", "True team-catalog tier")
        r = line(r, "Balls & equipment", f"={a('equip')}*1.5")
        r = line(r, "Production — marquee broadcast games", f"={a('marquee_fv')}*{a('marquee_rate_fv')}", "Broadcast-quality tier — the reel that gets you to ESPN conversations")
        r = line(r, "Production — standard streams", f"=(({games})-{a('marquee_fv')})*{a('std_rate_fv')}")
        r = line(r, "Player meals & transport", f"={a('px_fv')}")
        r = line(r, "Sponsor-funded player NIL/appearance pool", f"={a('nil_fv')}", "Legal under TSSAA NIL policy — no school marks, no coach facilitation")
        r = line(r, "Nationals — travel subsidies for invited teams", f"={a('natl_travel_fv')}")
        r = line(r, "Nationals — event operations", f"={a('natl_ops_fv')}")
        r = line(r, "Marketing & brand", f"={a('mktg_fv')}")
        r = line(r, "Insurance", f"={a('ins_fv')}")
        r = line(r, "Compliance", f"={a('compliance')}*1.5")
        r = line(r, "Tech & stats", f"={a('tech_fv')}")
        r = line(r, "Legal & accounting", f"={a('legal_fv')}")
        r = line(r, "Awards & trophies", f"={a('awards_fv')}")
    exp_last = r-1
    B.cell(row=r, column=1, value="Subtotal").font = BOLD
    c = B.cell(row=r, column=2, value=f"=SUM(B{exp_first}:B{exp_last})"); c.font = BOLD; c.number_format = CUR
    sub_row = r; r += 1
    B.cell(row=r, column=1, value="Contingency").font = BLACK
    c = B.cell(row=r, column=2, value=f"=B{sub_row}*{a('conting')}"); c.number_format = CUR
    r += 1
    B.cell(row=r, column=1, value="TOTAL EXPENSES").font = BOLD12
    c = B.cell(row=r, column=2, value=f"=B{sub_row}+B{r-1}"); c.font = BOLD12; c.number_format = CUR; c.fill = TOTFILL
    tot_exp = r; r += 2

    r = section(B, r, "REVENUE (base case — see Revenue Scenarios tab)", 4)
    rev_first = r
    if year1:
        r = line(r, "Tryout fees", f"={a('tryouts')}*{a('tryout_fee')}", "Credited toward season fee for drafted players — netted in the player-fee line note")
        r = line(r, "Player season fees", f"={a('players')}*{a('player_fee_y1')}", "Net of tryout credit; ~10x cheaper than an AAU summer")
        r = line(r, "Title sponsor", f"={a('title_y1')}")
        r = line(r, "Presenting sponsors", f"={a('pres_n_y1')}*{a('pres_p_y1')}")
        r = line(r, "Official category partners", f"={a('cat_n_y1')}*{a('cat_p_y1')}")
        r = line(r, "Team jersey partners", f"={a('jers_n_y1')}*{a('jers_p_y1')}")
        r = line(r, "Gate / tickets", f"={a('gate_y1')}")
        r = line(r, "Streaming rev share / league pass", f"={a('stream_y1')}")
        r = line(r, "Merchandise (net)", f"={a('merch_y1')}")
        r = line(r, "Donations & grants (501c3)", f"={a('grants_y1')}")
    else:
        r = line(r, "Player fees", f"={a('players')}*{a('player_fee_fv')}", "$0 — fully sponsored is the endgame")
        r = line(r, "Title sponsor", f"={a('title_fv')}")
        r = line(r, "All other sponsorship", f"={a('spons_other_fv')}", "See Sponsor Menu tab")
        r = line(r, "Gate / tickets", f"={a('gate_fv')}")
        r = line(r, "Media / streaming", f"={a('media_fv')}")
        r = line(r, "Merchandise (net)", f"={a('merch_fv')}")
        r = line(r, "Nationals entry fees", f"={a('natl_entry_fv')}")
        r = line(r, "Donations & grants", f"={a('grants_fv')}")
    rev_last = r-1
    B.cell(row=r, column=1, value="TOTAL REVENUE").font = BOLD12
    c = B.cell(row=r, column=2, value=f"=SUM(B{rev_first}:B{rev_last})"); c.font = BOLD12; c.number_format = CUR; c.fill = TOTFILL
    tot_rev = r; r += 2
    B.cell(row=r, column=1, value="NET SURPLUS / (FUNDING GAP)").font = BOLD12
    c = B.cell(row=r, column=2, value=f"=B{tot_rev}-B{tot_exp}"); c.font = BOLD12; c.number_format = CUR; c.fill = YELLOW
    B.cell(row=r, column=4, value="A gap here = the founding-capital raise you need (investors, founding partners, grants)").font = Font(name=ARIAL, size=9, color="595959")
    return B, tot_exp, tot_rev, rev_first, rev_last

Y1, y1_exp, y1_rev, y1_rf, y1_rl = build_budget("Year 1 Budget", True)
FV, fv_exp, fv_rev, _, _ = build_budget("Full Vision Budget", False)

# ---------------------------------------------------------------- REVENUE SCENARIOS
R = WB.create_sheet("Revenue Scenarios")
sheet_title(R, "Year 1 Revenue Scenarios", 5)
for col, w in zip("ABCDE", [40, 16, 16, 16, 60]):
    R.column_dimensions[col].width = w
r = 3
R.cell(row=r, column=1, value="Scenario multipliers (editable)").font = BOLD
r += 1
labels = ["Conservative", "Base", "Stretch"]
mults = [0.6, 1.0, 1.4]
mult_rows = {}
for lbl, m in zip(labels, mults):
    R.cell(row=r, column=1, value=f"{lbl} multiplier").font = BLACK
    c = R.cell(row=r, column=2, value=m); c.font = BLUE; c.number_format = '0.0'
    mult_rows[lbl] = r
    r += 1
R.cell(row=r, column=5, value="Multipliers apply to sponsorship, gate, streaming, merch & grants. Player/tryout fees held flat.").font = Font(name=ARIAL, size=9, color="595959")
r += 2
hdr = ["Line", "Conservative", "Base", "Stretch", "Notes"]
for i, h in enumerate(hdr, 1):
    c = R.cell(row=r, column=i, value=h); c.font = BOLD; c.fill = SECFILL
r += 1
scn_first = r
fee_lines = 2  # first two revenue rows are fee-based (flat)
for i, src_row in enumerate(range(y1_rf, y1_rl+1)):
    lbl_ref = f"='Year 1 Budget'!$A${src_row}"
    val_ref = f"'Year 1 Budget'!$B${src_row}"
    R.cell(row=r, column=1, value=lbl_ref).font = GREEN
    for j, lbl in enumerate(labels):
        mrow = mult_rows[lbl]
        if i < fee_lines:
            f = f"={val_ref}"
        else:
            f = f"={val_ref}*$B${mrow}"
        c = R.cell(row=r, column=2+j, value=f); c.font = BLACK; c.number_format = CUR
    for col in range(1,6): R.cell(row=r, column=col).border = THIN
    r += 1
scn_last = r-1
R.cell(row=r, column=1, value="TOTAL REVENUE").font = BOLD
for j in range(3):
    col = get_column_letter(2+j)
    c = R.cell(row=r, column=2+j, value=f"=SUM({col}{scn_first}:{col}{scn_last})")
    c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
tot_row = r; r += 1
R.cell(row=r, column=1, value="Total expenses (Year 1)").font = BLACK
for j in range(3):
    c = R.cell(row=r, column=2+j, value=f"='Year 1 Budget'!$B${y1_exp}")
    c.font = GREEN; c.number_format = CUR
exp_row = r; r += 1
R.cell(row=r, column=1, value="NET SURPLUS / (FUNDING GAP)").font = BOLD
for j in range(3):
    col = get_column_letter(2+j)
    c = R.cell(row=r, column=2+j, value=f"={col}{tot_row}-{col}{exp_row}")
    c.font = BOLD; c.number_format = CUR; c.fill = YELLOW

# ---------------------------------------------------------------- SPONSOR MENU
M = WB.create_sheet("Sponsor Menu")
sheet_title(M, "Sponsor Menu — sellable assets & Year 1 price points", 6)
for col, w in zip("ABCDEF", [42, 10, 14, 14, 16, 70]):
    M.column_dimensions[col].width = w
r = 3
hdr = ["Asset", "# Avail", "Price low", "Price high", "Potential (mid)", "What the sponsor gets"]
for i, h in enumerate(hdr, 1):
    c = M.cell(row=r, column=i, value=h); c.font = BOLD; c.fill = SECFILL
r += 1
menu_first = r
menu = [
    ("Title / naming rights ('The ___ League')", 1, 50000, 75000, "Name in league name, center court, all broadcasts, jerseys, every social post"),
    ("Presenting partner ('presented by')", 2, 20000, 30000, "Logo lockup on all media, in-arena signage, PA reads, category exclusivity"),
    ("Official ball / apparel / footwear partner", 3, 10000, 20000, "Product on court every game + retail activation rights"),
    ("Official category partner (bank, auto, health, QSR, drink)", 5, 7500, 15000, "Category exclusivity, court decal, stream ad slots, booth"),
    ("Team jersey front partner (per team)", 16, 4000, 6000, "Front-of-jersey patch on one team, in every game + highlight clip"),
    ("Court logo decal", 4, 5000, 7500, "On-court logo visible in every stream frame"),
    ("Combine title sponsor", 1, 10000, 15000, "Naming of the combine weekend, testing-station branding, content series"),
    ("Draft night title sponsor", 1, 10000, 15000, "Naming of draft night, stage branding, draft-cap co-brand"),
    ("Nationals invitational title (when launched)", 1, 25000, 50000, "Naming of the national event"),
    ("Halftime / in-game segment ('Play of the Night by ___')", 6, 2000, 4000, "Branded recurring content segment, Unrivaled 'Galaxy Game Winner' style"),
    ("Stream broadcast ad package (per session)", 8, 2500, 5000, "30-sec spots + lower-thirds across all streamed games"),
    ("Awards & MVP sponsor", 2, 4000, 6000, "Naming on MVP, All-League team, trophy presentations"),
    ("Scholarship / community fund partner", 3, 5000, 15000, "Funds player fee waivers — the anti-AAU story sponsors love; 501(c)(3) deductible"),
    ("Content series partner (mic'd up, all-access doc)", 3, 5000, 15000, "Branded episodic content — OTE/Unrivaled's actual growth engine"),
    ("Digital / website / app partner", 2, 3000, 6000, "Logo on stats hub, schedules, recruiting profiles"),
]
for asset, n, lo, hi, gets in menu:
    M.cell(row=r, column=1, value=asset).font = BLACK
    c = M.cell(row=r, column=2, value=n); c.font = BLUE; c.number_format = NUM
    c = M.cell(row=r, column=3, value=lo); c.font = BLUE; c.number_format = CUR
    c = M.cell(row=r, column=4, value=hi); c.font = BLUE; c.number_format = CUR
    c = M.cell(row=r, column=5, value=f"=B{r}*(C{r}+D{r})/2"); c.font = BLACK; c.number_format = CUR
    M.cell(row=r, column=6, value=gets).font = Font(name=ARIAL, size=9, color="595959")
    for col in range(1,7): M.cell(row=r, column=col).border = THIN
    r += 1
menu_last = r-1
M.cell(row=r, column=1, value="TOTAL MENU POTENTIAL (if fully sold, mid prices)").font = BOLD
c = M.cell(row=r, column=5, value=f"=SUM(E{menu_first}:E{menu_last})"); c.font = BOLD; c.number_format = CUR; c.fill = TOTFILL
r += 2
M.cell(row=r, column=1, value="Pricing method: $0.10-$2.00 per audience member x visibility multiplier (jersey 2-4x, PA/banner 1.5x), category exclusivity 2-5x premium — Prep Athletic Director model. Streaming impressions are the upsell lever.").font = Font(name=ARIAL, size=9, italic=True, color="595959")

WB.save("/home/claude/thsl/TESL_Budget_Model.xlsx")
print("saved")
